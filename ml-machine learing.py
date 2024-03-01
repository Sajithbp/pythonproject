from openpyxl import load_workbook

work_book = load_workbook("C:\\Users\\USER\\Documents\\feb 121.xlsx")

sheet1 = work_book["Sheet1"]
data_sheet1 = []

for row in sheet1.iter_rows(values_only=True):
    data_row = []
    for cell in row:
        data_row.append(cell)
    data_sheet1.append(data_row)

sheet2 = work_book["Sheet2"]
data_sheet2 = []

for row in sheet2.iter_rows(values_only=True):
    data_row = []
    for cell in row:
        data_row.append(cell)
    data_sheet2.append(data_row)

print("Data from sheet1:")
for row in data_sheet1:
    print(row)

print("Data from sheet2:")
for row in data_sheet2:
    print(row)
