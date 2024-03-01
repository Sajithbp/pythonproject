"""Methodover loading"""
#
# class overloading():
#     def add(self,a,b):
#         return a+b
#
#     def add(self,a,b,c):
#         return a+b+c



from openpyxl import load_workbook

work_book = load_workbook("C:\\Users\\USER\\Documents\\feb 121.xlsx")
sheet1 = work_book["sheet1"]
data_sheet1 = []

for row in sheet1.iter_rows(values_only=True):
    data_row = []
    for cell in row:
        data_row.append(cell)
    data_sheet1.append(data_row)

sheet2 = work_book["sheet2"]
data_sheet2 = []

for row in sheet2.iter_rows(values_only=True):
    data_row = []
    for cell in row:
        data_row.append(cell)
    data_sheet2.append(data_row)

for row in data_sheet1:
    print(row)

for row in data_sheet2:
    print(row)

#
#
# opps=overloading()
# # result=opps.add(2,3,4)
# #
# # print(result)
#
#
#
# result1=opps.add(2,3)
# print("result1:",result1)


# import openpyxl
# book=openpyxl.load_workbook(r"C:\Users\USER\Desktop\Copy of Calculate Sales Commission(1).xlsx")
# sheet=book.active
#
#
# for row in sheet.iter_rows(values_only=True):
#     for cell in row:
#         print(cell,end="   \t")
#     print()



import pandas as pd
excel=pd.read_excel(r"C:\Users\USER\Desktop\Copy of Calculate Sales Commission(1).xlsx",nrows=15)
print(excel.head())
print(excel)